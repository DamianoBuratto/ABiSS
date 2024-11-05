#!/bin/python3
# The program has been made to analyze the data of ABiSS program.
# It returns a plot of the BE with STD for every mutation (black WT, blue accepted, red declined)
# Use it inside the RUN folder
# Use conda activate gmxMMPBSA
# e.g. python Data_analysis.py -i conf_RUN1.out -l Config0/RESULTS/MoleculesResults.dat Mutant?/RESULTS/MoleculesResults.dat  Mutant??/RESULTS/MoleculesResults.dat -u "KCal/mol"
import argparse
import openpyxl
from openpyxl.styles import Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import ScatterChart, Reference, Series

import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import pandas as pd
import os
import glob
import json

aa_notation_3to1 = {
    "GLY": "G",
    "ALA": "A",
    "VAL": "V",
    "LEU": "L",
    "ILE": "I",
    "MET": "M",
    "PHE": "F",
    "TRP": "W",
    "PRO": "P",

    "SER": "S",
    "THR": "T",
    "CYS": "C",
    "TYR": "Y",
    "ASN": "N",
    "GLN": "Q",

    "ASP": "D",
    "GLU": "E",

    "LYS": "K",
    "ARG": "R",
    "HIS": "H"
}


def load_molecules_results(input_file_list, convertion, energy_ndx=1):
    # Open the 'Config0/RESULTS/MoleculesResults.dat' data file and read the data as a list of row. Compute avg and std
    print(f"load_molecules_results from {input_file_list}")
    average_list = []
    std_list = []
    first_file_length = 0
    conv_coeff = 1
    if convertion:
        conv_coeff = 0.24
    for input_file in input_file_list:
        with open(input_file) as f:
            rows = [line.split() for line in f]
        if not first_file_length:
            first_file_length = len(rows)
            print("first_file_length=" + str(first_file_length))
        if len(rows) != first_file_length:
            print(input_file + " (rows=" + str(len(rows)) + "): NOT FINISHED")
            continue
        binding_eng = [float(item[energy_ndx]) for item in rows[1:]]
        # create a numpy array
        binding_eng_array = np.array(binding_eng)
        average = np.mean(binding_eng_array)
        average = np.round(average, 1)
        std = np.std(binding_eng_array)
        std = np.round(std, 1)
        average_list.append(average * conv_coeff)
        std_list.append(std * conv_coeff)
        # print(input_file + ": " + str(binding_eng))
        print(input_file + ": " + str(average) + " +/- " + str(std))
    return [average_list, std_list]


def read_confout_file(input_file):
    print("read_confout_file.. ")
    # Open the data file and read the data into a list of row
    with open(input_file) as f:
        rows = [line.split() for line in f if len(line.strip()) > 0]

    # There is an error on old file that cause the dy to be in the same column with kj/mol.
    # I am going to strip kj/mol from every value. Similarly for "(" and ")"
    # I also fix the error that makes a new column with ")" from the second row on
    for line in rows:
        while line.count(')'):
            line.remove(')')
        for ndx, value in enumerate(line):
            if "kJ/mol" in value:
                line[ndx] = value.replace("kJ/mol", "")
                value = line[ndx]
            if "(" in value:
                line[ndx] = value.replace("(", "")
                value = line[ndx]
            if ")" in value:
                line[ndx] = value.replace(")", "")
            # if ndx == first_row_len or ndx == first_row_len + 2:
            #     line[ndx] = float(value)

    # Delete the last row if it is incomplete
    print("len(rows[-1]): " + str(len(rows[-1])))
    try:
        print("len(rows[2]): " + str(len(rows[2])))
        while len(rows[-1]) < len(rows[2]):
            print("Last Mutation wasn't completed.. ")
            print("delete: " + str(rows[-1]))
            del rows[-1]
            print("new last row: " + str(rows[-1]))
    except:
        print("Probably only WT..")

    return rows


def print_json_and_csv(confout_input_file_rows, output_file_name, convertion, run_number,
                       binding_eng_2, std_binding_eng_2):
    print(f'=> print_json_and_csv..')
    first_row_len = len(confout_input_file_rows[0])
    conv_coeff = 1
    if convertion:
        conv_coeff = 0.24
    # system_name -> WT, M1, M2, ...
    system_name = [item[0] for item in confout_input_file_rows[1:]]
    # system_sequence -> [MET, MET, TRP ...], [...]
    system_sequence = [item[1:first_row_len] for item in confout_input_file_rows[1:]]
    binding_eng = [float(item[first_row_len]) * conv_coeff for item in confout_input_file_rows[1:]]
    delta_binding_eng = [float(item[first_row_len]) * conv_coeff - float(confout_input_file_rows[1][first_row_len])
                         for item in confout_input_file_rows[1:]]
    std_binding_eng = [float(item[first_row_len + 2]) * conv_coeff for item in confout_input_file_rows[1:]]
    delta_binding_eng2 = np.round([item - binding_eng_2[0] for item in binding_eng_2], 1)
    delta_binding_eng2 = np.round(delta_binding_eng2, 1)

    print(f'\tfirst line example for RUN{run_number}:\n\t\t{system_name[0]} '
          f'{system_sequence[0]} {binding_eng[0]} {std_binding_eng[0]} {delta_binding_eng[0]} {delta_binding_eng2[0]}')
    # Make a dictionary with the sequences sampled during the run

    run_dic = {f'RUN{run_number}': []}
    run_list = []
    # for name, seq, BE, STD in zip(system_name, system_sequence, binding_eng, std_binding_eng):
    for name, seq, condBE, confSTD, confDBE, BE, STD, DBE in zip(system_name, system_sequence,
                                                                 binding_eng, std_binding_eng, delta_binding_eng,
                                                                 binding_eng_2, std_binding_eng_2,
                                                                 delta_binding_eng2):
        run_list.append({'run': f'RUN{run_number}', 'mutation': name, 'sequence': seq,
                         'binding_energy': f'{condBE:.2f}', 'std': f'{confSTD:.2f}',
                         'delta_binding_energy': f'{confDBE:.2f}', 'binding_energy_2s': f'{BE:.2f}',
                         'std_2s': f'{STD:.2f}', 'delta_binding_energy_2s': f'{DBE:.2f}'})
        run_dic[f'RUN{run_number}'].append({'mutation': name, 'sequence': seq, 'binding_energy': condBE, 'std': confSTD,
                                            'delta_binding_energy': f'{confDBE:.2f}', 'binding_energy_2s': f'{BE:.2f}',
                                            'std_2s': f'{STD:.2f}', 'delta_binding_energy_2s': f'{DBE:.2f}'})
    with open(output_file_name + ".json", 'w') as json_file:
        json.dump(run_dic, json_file, indent=4)
    with open(output_file_name + "_2.json", 'w') as json_file:
        json.dump(run_list, json_file, indent=4)
    df_run_list = pd.DataFrame(run_list)
    df_run_list.to_csv(output_file_name + ".csv", index=False)  # index=false means do not write rox indexes

    return run_list, run_dic


def make_table_excel_and_plot(confout_input_file_rows, output_file_name, excel_file_name, be_unit,
                              convertion, print_mutation, binding_eng2, std_binding_eng2):
    """
    system_name
    system_sequence
    binding_eng
    delta_binding_eng
    std_binding_eng
    """
    print("make_table_excel_and_plot.. ")
    # input_file_rows = read_confout_file(input_file)
    # first_row_len -> ResNum   29:H    30:H    31:H
    first_row_len = len(confout_input_file_rows[0])
    # second_row_len = len(confout_input_file_rows[1])
    conv_coeff = 1
    if convertion:
        conv_coeff = 0.24
    # system_name -> WT, M1, M2, ...
    system_name = [item[0] for item in confout_input_file_rows[1:]]
    # system_sequence -> MET     MET     TRP  ...
    system_sequence = [item[1:first_row_len] for item in confout_input_file_rows[1:]]
    binding_eng = [float(item[first_row_len]) * conv_coeff for item in confout_input_file_rows[1:]]
    delta_binding_eng = [float(item[first_row_len]) * conv_coeff - float(confout_input_file_rows[1][first_row_len])
                         for item in confout_input_file_rows[1:]]
    std_binding_eng = [float(item[first_row_len + 2]) * conv_coeff for item in confout_input_file_rows[1:]]

    # Make a dictionary with the sequences sampled during the run
    run_dic = {}
    run_dic[f'RUN{run_number}'] = []
    for name, seq, BE, STD in zip(system_name, system_sequence, binding_eng, std_binding_eng):
        run_dic[f'RUN{run_number}'].append({'mutation': name, 'sequence': seq, 'binding_energy': BE, 'std': STD})
    with open(input_file[:-4] + ".json", 'w') as json_f:
        json.dump(run_dic, json_f, indent=4)

    # Different colors for WT (black), ACCEPTED (blue) and DECLINED (red)
    def color_legend(result):
        if result == "**ACCEPTED**":
            return "blue"
        elif result == "**DECLINED**":
            return "red"
        else:
            return "green"

    accepted_declined = [color_legend(item[-1]) for item in confout_input_file_rows[2:]]
    accepted_declined.insert(0, "black")

    # Create a new Excel workbook and select the first sheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the data to the worksheet
    worksheet.cell(row=1, column=first_row_len + 3, value=be_unit)
    # row_index will start from 0, but worksheet start from 1
    for row_index, row in enumerate(confout_input_file_rows):
        max_column_index = 0
        for column_index, value in enumerate(row):
            if row_index > 0 and (column_index == first_row_len or column_index == first_row_len + 2):
                # These will be Avg_BE and Std_BE
                worksheet.cell(row=row_index + 1, column=column_index + 1, value=float(value) * conv_coeff)
                continue
            worksheet.cell(row=row_index + 1, column=column_index + 1, value=value)
            max_column_index = column_index
        if row_index == 1:
            max_column_index += 2
        if row_index >= 1:
            worksheet.cell(row=row_index + 1, column=max_column_index + 2, value=binding_eng[row_index - 1])
            worksheet.cell(row=row_index + 1, column=max_column_index + 3, value=std_binding_eng[row_index - 1])
            worksheet.cell(row=row_index + 1, column=max_column_index + 4, value=delta_binding_eng[row_index - 1])

    # Add formatting to the first row and column labels
    font_blue = Font(color="0000FF")
    for row in worksheet.rows:
        row[0].font = font_blue

    for cell in worksheet.columns:
        cell[0].font = font_blue

    # Add conditional formatting to cells
    font_red = Font(color="FF0000")
    font_green = Font(color="00FF00")
    last_accepted_row = 1
    for row_index, row in enumerate(confout_input_file_rows):
        if row_index < 2: continue
        system_name[row_index - 1] = "- " + system_name[row_index - 1]
        for column_index, value in enumerate(row):
            if "ACCEPTED" in value:
                worksheet.cell(row=row_index + 1, column=column_index + 1).font = font_green
                last_accepted_row = row_index
            if column_index == 0 or column_index >= first_row_len: continue

            # save the name if it's different from the starting configuration
            if row_index >= 2 and value != confout_input_file_rows[1][column_index]:
                worksheet.cell(row=row_index + 1, column=column_index + 1).font = font_red
                # print(str(confout_input_file_rows[1][column_index]) + '\t' + confout_input_file_rows[0][column_index] + '\t' + str(value))
                # system_name[row_index - 1] = confout_input_file_rows[1][column_index] + confout_input_file_rows[0][column_index][:-2] + value \
                #                             + "(" + confout_input_file_rows[0][column_index][-1:] + ") " + system_name[row_index - 1]
                system_name[row_index - 1] = aa_notation_3to1[confout_input_file_rows[1][column_index]] + \
                                             confout_input_file_rows[0][column_index][:-2] + \
                                             aa_notation_3to1[value] + "(" + confout_input_file_rows[0][column_index][
                                                                             -1:] + ") " + \
                                             system_name[row_index - 1]

            # save the name if it's different from the last accepted but equal to the starting one (WT)
            elif row_index >= 2 and value != confout_input_file_rows[last_accepted_row][column_index]:
                worksheet.cell(row=row_index + 1, column=column_index + 1).font = font_red
                # print(str(confout_input_file_rows[1][column_index]) + '\t' + confout_input_file_rows[0][column_index] + '\t' + str(value))
                # system_name[row_index - 1] = "(" + confout_input_file_rows[last_accepted_row][column_index] + ")" + \
                #                              confout_input_file_rows[0][column_index][:-2] + value + "(" + confout_input_file_rows[0][column_index][-1:] \
                #                              + ") " + system_name[row_index - 1]
                system_name[row_index - 1] = "(" + aa_notation_3to1[
                    confout_input_file_rows[last_accepted_row][column_index]] + ")" + \
                                             confout_input_file_rows[0][column_index][:-2] + aa_notation_3to1[
                                                 value] + "(" + \
                                             confout_input_file_rows[0][column_index][-1:] + ") " + system_name[
                                                 row_index - 1]
                # formula = f"NOT({cell.coordinate}={cell.column_letter}$2)"
                # rule = FormulaRule(formula=formula, font=Font(color="FF0000"))
                # worksheet.conditional_formatting.add(cell.coordinate, rule)

    # Create a scatter chart
    chart = ScatterChart()
    x_values = Reference(worksheet, min_col=1, min_row=2, max_row=len(confout_input_file_rows))
    y_values = Reference(worksheet, min_col=len(confout_input_file_rows[1]) - 2, min_row=2,
                         max_row=len(confout_input_file_rows))
    y_error = Reference(worksheet, min_col=len(confout_input_file_rows[1]), min_row=2,
                        max_row=len(confout_input_file_rows))
    series = Series(y_values, x_values, y_error)
    chart.series.append(series)
    chart.title = "Scatter Plot"
    chart.x_axis.title = "X Values"
    chart.y_axis.title = "Y Values"
    worksheet.add_chart(chart, "AA1")

    # Save the Excel file
    workbook.save(excel_file_name + ".xlsx")

    # Make a plot with the data
    # make_plot(x_value=system_name, y_value=binding_eng, y_err=std_binding_eng, plot_name=input_file[:-4])
    new_system_name = []
    previous_word = []
    for i, name in enumerate(system_name):
        a_word = name.split()
        if print_mutation:
            new_name = ' '.join(word for word in a_word if word == "-" or word not in previous_word)
        else:
            new_name = a_word[-1]
        new_system_name.append(new_name)
        print(
            'a_word=\t\t' + str(a_word) + '\nprevious_word=\t' + str(previous_word) + '\nnew_name=\t' + new_name + '\n')
        if accepted_declined[i] == 'blue':
            previous_word = system_name[i].split()

    print('system_name[' + str(len(system_name)) + ']= ' + str(system_name) +
          '\nx_value[' + str(len(new_system_name)) + ']= ' + str(new_system_name) +
          '\ny_value[' + str(len(delta_binding_eng)) + ']= ' + str(list(delta_binding_eng)) +
          '\ny_err[' + str(len(std_binding_eng)) + ']= ' + str(std_binding_eng) +
          '\ncolor_legend[' + str(len(accepted_declined)) + ']= ' + str(accepted_declined) +
          '\nbe_unit= \"' + str(be_unit) + '\"')
    with open(output_file_name + ".dat", 'w') as f:
        print('x_value= ' + str(new_system_name) +
              "\ny_value= " + str(list(delta_binding_eng)) +
              "\ny_err= " + str(std_binding_eng) +
              "\ncolor_legend= " + str(accepted_declined) +
              "\nbe_unit= \"" + str(be_unit) + "\"", file=f)

    print('> RUNNING make_plot..')
    return [delta_binding_eng, std_binding_eng, new_system_name, accepted_declined]
    # make_plot_png(x_value=new_system_name, y_value=delta_binding_eng, y_err=std_binding_eng,
    #              color_legend=accepted_declined, plot_name=input_file[:-4], be_unit=be_unit,
    #              print_mutation=print_mutation)


def make_plot_png(x_value, y_value, y_err, color_legend, plot_name, be_unit, print_mutation):
    # Create figue and axis
    # fig, ax = plt.subplots()
    skip_xticks = 1
    size = [100 for item in range(1, len(x_value))]
    size.insert(0, 150)
    fig = plt.figure(figsize=(18, 12))
    plt.scatter(x_value, y_value, c=color_legend, s=size)
    x_accepted = [x_value[i] for i in range(0, len(x_value)) if color_legend[i] == "blue" or color_legend[i] == "black"]
    y_accepted = [y_value[i] for i in range(0, len(y_value)) if color_legend[i] == "blue" or color_legend[i] == "black"]
    plt.plot(x_accepted, y_accepted, c="blue", linestyle='--', linewidth=2)

    # Plot the data with error bars
    # ax.errorbar(x_value, y_value, yerr=y_err, fmt='o', capsize=5)
    plt.errorbar(x_value, y_value, yerr=y_err, capsize=15, fmt='none', ecolor="grey", alpha=0.6, ls="-")

    # Add title and labels and grid
    # ax.set_title(plot_name)
    # ax.set_ylabel('Binding Energy [kJ/mol]')
    plt.title(plot_name)
    plt.ylabel('Binding Energy [' + be_unit + ']', fontsize=45)
    plt.yticks(fontsize=30)q
    if print_mutation or len(x_value) > 15:
        if len(x_value) > 15:
            plt.xticks(rotation=80, fontsize=30)
        else:
            plt.xticks(rotation=80, fontsize=30)
    else:
        plt.xticks(fontsize=30)
    ax = plt.gca()
    # ax.set_xticklabels(x_value)
    ax.xaxis.set_major_locator(ticker.MultipleLocator(base=skip_xticks))
    for i, label in enumerate(ax.get_xticklabels()):
        # label_x_position = int(label.get_position()[0])
        label_x_position = i - 1
        if label_x_position < 0: continue
        if label_x_position >= len(color_legend): break
        print('i=' + str(i) + '\t label=' + str(label) + '\t (i-1)*skip_xticks=' + str((i - 1) * skip_xticks) +
              '\t label.get_position()=' + str(label.get_position()) +
              '\t color_legend[(i-1)*skip_xticks]=' + color_legend[(i - 1) * skip_xticks] +
              '\t label_x_position=' + str(label_x_position) +
              '\t color_legend[label_x_position]=' + color_legend[label_x_position])
        label.set_color(color_legend[label_x_position])

    plt.grid(linestyle='--', linewidth=0.5)

    # Save table as PNG
    plt.savefig(plot_name + '.png', bbox_inches='tight', dpi=300)

    # Show the plot
    # plt.show(block=False)
    # plt.show()

    # This code reads the file "data.txt", extracts the labels and data, transposes the data to
    # create columns, creates a table using the `ax.table()` function from the matplotlib library,
    # sets the font size and color for the labels and the cells, and saves the table as a PNG image
    # using the `plt.savefig()` function.


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Data analysis for ABiSS program')
    parser.add_argument('-f', '--input_folder_list', type=str, nargs='+', default="RUN1",
                        help='List of folder where the conf input file (-i opt) can be found (Default: RUN1)')
    parser.add_argument('-i', '--input_file', type=str, default="conf_RUN1.out",
                        help='Input file name with residues list and BE (Default: conf_RUN1.out)')
    parser.add_argument('-l', '--input_file_list', type=str, nargs='+', default=None,
                        help='List of MoleculesResults.dat input file name (Default: None)')
    parser.add_argument('-o', '--output_name', type=str, default="data_excel",
                        help='Name of the excel output file (Default: data_excel)')
    parser.add_argument('-u', '--be_unit', type=str, default="kJ/mol",
                        help='Unit used to measure the binding energy. It will be used as label (Default: kJ/mol)')
    parser.add_argument('-m', '--print_mutation', action='store_false', default=True,
                        help='Print the mutation in the x_labels (Default:True)')
    parser.add_argument('-r', '--run_number', type=int, default=1,
                        help='Number of the RUN that you are analysing (Default:1)')
    parser.add_argument('-c', '--convertion', action='store_true', default=False,
                        help='Convert the BE from KJ/mol (old MMPBSA) to KCal/mol. Converting coefficient: 0.24)')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose output')
    args = parser.parse_args()

    # VERBOSE NOT IMPLEMENTED YET
    verbose = args.verbose
    input_file_name = args.input_file
    base_directory = os.getcwd()
    input_file_list_glob = []
    for file in args.input_file_list:
        # print(f"file: {file}->{glob.glob(file)}")
        input_file_list_glob.extend(glob.glob(file))
    print(f"input_file_list_glob: {input_file_list_glob}")

    # if args.input_file_list:
    #     load_molecules_results(input_file_list=args.input_file_list)
    #     exit()
    confout_file_rows = read_confout_file(input_file_name)
    binding_eng2 = 0
    std_binding_eng2 = 0
    if input_file_list_glob:
        binding_eng2, std_binding_eng2 = load_molecules_results(input_file_list_glob, args.convertion, energy_ndx=10)

    [d_BE, std_BE, new_names, acc_dec] = make_table_excel_and_plot(confout_input_file_rows=confout_file_rows,
                                                                   input_file_list=input_file_list_glob,
                                                                   excel_file_name=args.output_name,
                                                                   be_unit=args.be_unit,
                                                                   convertion=args.convertion,
                                                                   print_mutation=args.print_mutation,
                                                                   run_number=args.run_number,
                                                                   binding_eng2=binding_eng2,
                                                                   std_binding_eng2=std_binding_eng2)

    make_plot_png(x_value=new_names, y_value=d_BE, y_err=std_BE, color_legend=acc_dec, plot_name=input_file_name[:-4],
                  be_unit=args.be_unit, print_mutation=args.print_mutation)
